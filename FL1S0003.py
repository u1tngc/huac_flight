#PGM-ID:FL1S0003
#PGM-NAME:FL単座チェック表EXCEL出力サブ
#最終更新日:2026/08/09

import os
import shutil
import tempfile
from datetime import datetime, timezone, timedelta
from urllib.parse import quote
from flask import Response
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import FL0S002D
import FL0S099D
import FL1S0001

#テンプレ格納先（FL1S0003.pyと同階層の template フォルダ）
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "template", "合宿前資料_テンプレ.xlsx")

#文字色（ＮＧ:赤／ＯＫ:青）とＮＧ時のセル塗りつぶし色（灰）
COLOR_NG = "FFFF0000"
COLOR_OK = "FF0000FF"
FILL_NG = PatternFill("solid", fgColor="FFD9D9D9")

#EXCELのMIMEタイプ
EXCEL_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

DATA_START_ROW = 3          #明細開始行
COL_NAME = "A"              #学生氏名
COL_KINKYU = ["B", "C", "D", "E"]            #サブＧ／索切れ／失速／スピン
COL_SOLO = ["F", "G", "H", "I", "J"]         #ノーダイブ／横風着陸／オーバーヘッド／失速／口述
COL_SOLO_HANTEI = "K"
COL_23 = ["L", "M", "N"]                     #Ａ章／Ｂ章／２３移行口述
COL_23_GAKKA = "O"
COL_23_HANTEI = "P"
COL_DIS_ARCUS = "Q"
COL_DIS_SHIKAKU = "R"
COL_DIS_HANTEI = "S"

def _fmt_ymd(value):
    """日付をyyyy/mm/dd形式へ。未実施はそのまま、判定対象外は-を返す"""
    if value == "判定対象外":
        return "-"
    if value == "未実施":
        return value
    for fmt in ("%Y/%m/%d", "%y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y/%m/%d")
        except ValueError:
            pass
    return value

def _set_cell(ws, col, row, value, ng_flg, fill_flg=1):
    """セルへ値を設定
       ng_flg   1:赤字 2:青字 それ以外:テンプレのまま
       fill_flg 1:ＮＧ時に灰色で塗る 0:塗らない（判定列で使用）"""
    cell = ws[f"{col}{row}"]
    cell.value = value
    if ng_flg == 1:
        cell.font = Font(name=cell.font.name, size=cell.font.size,
                         bold=cell.font.bold, color=COLOR_NG)
        if fill_flg == 1:
            cell.fill = FILL_NG
    elif ng_flg == 2:
        cell.font = Font(name=cell.font.name, size=cell.font.size,
                         bold=cell.font.bold, color=COLOR_OK)

def make_SoloChk_excel():
    """単座チェック表を作成し、(バイト列, ファイル名) を返す"""
    JST = timezone(timedelta(hours=9))
    now = datetime.now(JST)

    #全学生分のデータを先読みしてキャッシュ（DB接続を4回に抑える）
    FL0S002D.load_cacheSolo()
    FL0S099D.load_cacheGakusei()
    FL0S099D.load_cacheTest()

    #テンプレを一時ファイルへコピー（テンプレ本体は編集しない）
    tmp_dir = tempfile.mkdtemp()
    tmp_path = os.path.join(tmp_dir, "work.xlsx")
    try:
        shutil.copy2(TEMPLATE_PATH, tmp_path)
        wb = load_workbook(tmp_path)
        ws = wb.active

        gakusei_list = FL1S0001.get_gakuseiInfo01()
        row = DATA_START_ROW
        for gakusei in gakusei_list:
            name, r0, r1, r2, r3 = FL1S0001.get_SoloChk(gakusei[0])

            #学生氏名
            _set_cell(ws, COL_NAME, row, name, 0)

            #①緊急処置課目
            #  6ヵ月以内 :[1]に実施日  6ヵ月超過:[1]は未実施のまま[2]に実施日
            #  データ無し :[1]は未実施 [2]は空
            for ix1 in range(len(COL_KINKYU)):
                value = r0[ix1][1]
                if value == "未実施" and r0[ix1][2]:
                    value = r0[ix1][2]
                _set_cell(ws, COL_KINKYU[ix1], row, _fmt_ymd(value), r0[ix1][0])

            #②単座要件
            for ix1 in range(len(COL_SOLO)):
                _set_cell(ws, COL_SOLO[ix1], row, _fmt_ymd(r1[ix1][1]), r1[ix1][0])
            if r1[5] == 1:
                _set_cell(ws, COL_SOLO_HANTEI, row, "ＮＧ", 1, 0)
            else:
                _set_cell(ws, COL_SOLO_HANTEI, row, "ＯＫ", 2)

            #③２３要件
            for ix1 in range(len(COL_23)):
                _set_cell(ws, COL_23[ix1], row, _fmt_ymd(r2[ix1][1]), r2[ix1][0])
            if r2[3][0] == 1:
                _set_cell(ws, COL_23_GAKKA, row, "ＮＧ", 1)
            else:
                _set_cell(ws, COL_23_GAKKA, row, "ＯＫ", 0)
            if r2[4] == 1:
                _set_cell(ws, COL_23_HANTEI, row, "ＮＧ", 1, 0)
            else:
                _set_cell(ws, COL_23_HANTEI, row, "ＯＫ", 2)

            #④ディスカス要件
            _set_cell(ws, COL_DIS_ARCUS, row, _fmt_ymd(r3[0][1]), r3[0][0])
            if r3[2][0] == 1:
                _set_cell(ws, COL_DIS_SHIKAKU, row, "未取得", 1)
            else:
                _set_cell(ws, COL_DIS_SHIKAKU, row, "取得済", 0)
            if r3[3] == 1:
                _set_cell(ws, COL_DIS_HANTEI, row, "ＮＧ", 1, 0)
            else:
                _set_cell(ws, COL_DIS_HANTEI, row, "ＯＫ", 2)

            row = row + 1

        wb.save(tmp_path)
        with open(tmp_path, "rb") as f:
            data = f.read()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        FL0S002D.clear_cacheSolo()
        FL0S099D.clear_cacheGakusei()
        FL0S099D.clear_cacheTest()

    filename = f"単座チェック表_{now.strftime('%Y%m%d%H%M')}.xlsx"
    return data, filename

def make_SoloChk_excelResponse():
    """単座チェック表をダウンロード用のResponseとして返す"""
    data, filename = make_SoloChk_excel()
    return Response(
        data,
        mimetype=EXCEL_MIMETYPE,
        headers={"Content-Disposition":
                 f"attachment; filename*=UTF-8''{quote(filename)}"}
    )